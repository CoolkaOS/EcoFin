import openpyxl
import wr
from openpyxl.utils import get_column_letter as let


def totable():
    wb = openpyxl.Workbook()
    ws = wb.active
    results = wr.read_results()
    ids = list(results.keys())
    al = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    for row in ws['A1:ZZ100']:
        for cell in row:
            cell.alignment = al
    ws.column_dimensions['A'].width = 18
    for i in range(2, 1000):
        ws.column_dimensions[let(i)].width = 5
    i = 0
    for id in ids:
        ws.cell(row=2 + 3 * i, column=1, value=results[ids[i]][1])
        ws.cell(row=3 + 3 * i, column=1, value=ids[i])
        i += 1
    index = 0
    for id in ids:
        le = len(results[id][4])
        for pr in results[id][4].items():
            ws.cell(row=2 + index * 3,
                    column=3 + list(results[id][4].items()).index(pr),
                    value=pr[0])
            ws.cell(row=3 + index * 3,
                    column=3 +list(results[id][4].items()).index(pr),
                    value=pr[1])
        ws.cell(row=3 + index * 3,
                column=2,
                value=sum(results[id][4][pr] for pr in results[id][4]))
        index += 1

    wb.save('res.xlsx')
